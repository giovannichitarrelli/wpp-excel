from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.common.keys import Keys
import pyperclip
import os
import openpyxl
from tkinter import *
from openpyxl import Workbook
import xlsxwriter 


def configs_web():
    global driver
    ###### Tira os erros de log no terminal
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-logging")
    options.add_argument("--log-level=3")
    ######
    dir_path = os.getcwd()
    profile = os.path.join(dir_path, "profile", "wpp")
    options.add_argument(r"user-data-dir={}".format(profile))
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=options)
    driver.maximize_window
    driver.get( "https://web.whatsapp.com/")

    while len(driver.find_elements(By.ID, "side")) < 1:
        time.sleep(1)
    time.sleep(2)


def download_wpp():
    configs_web()
    pesquisa = "#121"
    driver.find_element(By.XPATH, "//*[@id='side']/div[1]/div/div/div[2]/div/div[1]/p").send_keys(pesquisa, Keys.ENTER)
    time.sleep(1)
    info_grupo = driver.find_element(By.XPATH, "//*[@id='main']/header/div[2]/div[2]/span")
    info_grupo.click()
    time.sleep(1)

    numero = driver.find_element(By.XPATH, "//*[@id='main']/header/div[2]/div[2]/span").text
    workbook = xlsxwriter.Workbook('results.xlsx') 
    worksheet = workbook.add_worksheet() 
    row = 0
    column = 0
    itens = numero.split(',')
    for i in range(len(itens) - 1): 
        worksheet.write(row, column, itens[i]) 
        row += 1
    time.sleep(1)
    workbook.close() 
    driver.close()
    print("tudo certo com o arquivo")
    time.sleep(1)
def send_msg_wpp():
    configs_web()
    global count_sucedidos
    global count_inexistentes
    global count

    count_sucedidos = 0
    count_inexistentes = 0
    count = 0

    contatos = openpyxl.load_workbook("results.xlsx")
    sheet = contatos.active 
    num_contatos = sheet.max_row

    for contato in range(1, num_contatos + 1):
        number = sheet.cell(row=contato, column=1).value
        count = count + 1
        print(f'Valor na contagem {count}: {number}')

        driver.get(f"https://web.whatsapp.com/send?phone={number}")
        while len(driver.find_elements(By.ID, "side")) < 1:
            time.sleep(1)
        time.sleep(2)

        if len(driver.find_elements(By.XPATH, '//*[@id="app"]/div/span[2]/div/span/div/div/div/div')) < 1: 
            mensagem = """Olá, msg do bot teste!"""
            pyperclip.copy(mensagem) 
            driver.find_element(By.XPATH, '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[1]').send_keys(Keys.CONTROL + 'v')
            time.sleep(2)
            driver.find_element(By.XPATH, '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[1]').send_keys(Keys.ENTER)
            time.sleep(2)

            count_sucedidos = count_sucedidos + 1
            print(f'Mensagem enviada: {number}')
            print( 50 * '-')    
            time.sleep(1)
        else:
            count_inexistentes = count_inexistentes + 1
            print(f'Mensagem NÃO enviada: {number}')
            print( 50 * '-')   
            time.sleep(1)
        print("esperando 3s...")
        time.sleep(3)

    print(f"contatos NÃO sucedidos: {count_inexistentes}")
    print(f"contatos bem sucedidos: {count_sucedidos}")
    print(f"contatos Total: {count}")
    driver.close()

janela = Tk()
janela.title("Whatsapp")

texto_orientacao = Label(janela, text="Inicie o programa clicando no botão")
texto_orientacao.grid(column=0, row=0, padx=70, pady=10)


btn_send = Button(janela, text="Enviar mensagens" , command=send_msg_wpp)
btn_send.grid(column=0, row=2,padx=70, pady=10)


btn_down = Button(janela, text="Buscar contatos" , command=download_wpp) 
btn_down.grid(column=0, row=3,padx=70, pady=10)

janela.mainloop()